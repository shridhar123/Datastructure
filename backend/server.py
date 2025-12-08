from fastapi import FastAPI, APIRouter, File, UploadFile, Form, HTTPException
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import shutil
from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType
import pandas as pd
import openpyxl
import json

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create uploads directory
UPLOADS_DIR = ROOT_DIR / 'uploads'
UPLOADS_DIR.mkdir(exist_ok=True)

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# LLM Configuration
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY')
if not EMERGENT_LLM_KEY:
    logger.warning("EMERGENT_LLM_KEY not set in environment variables. LLM features will not work.")

# Define Models
class PDFUploadResponse(BaseModel):
    id: str
    filename: str
    file_path: str
    uploaded_at: str
    status: str

class ConversionRequest(BaseModel):
    file_id: str
    prompt: str

class ConversionResponse(BaseModel):
    id: str
    file_id: str
    excel_path: str
    prompt: str
    status: str
    created_at: str

class ColumnMapping(BaseModel):
    client_column: str
    icyte_column: str
    operation: Optional[str] = None  # e.g., "multiply:2", "add:10", "subtract:5"

class MappingTemplate(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    mappings: List[ColumnMapping]
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ReconciliationConfig(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_file_id: str
    icyte_file_id: str
    client_sheet: str
    icyte_sheet: str
    client_unique_key: str  # Column to use as unique identifier in client file
    icyte_unique_key: str   # Column to use as unique identifier in ICyte file
    mappings: List[ColumnMapping]
    template_id: Optional[str] = None  # Reference to the template used
    template_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class ReconciliationReport(BaseModel):
    id: str
    config_id: str
    total_records: int
    matched_records: int
    variances: int
    exceptions: List[Dict[str, Any]]
    summary: Dict[str, Any]
    created_at: str

# Routes
@api_router.get("/")
async def root():
    return {"message": "Reconciliation API"}

@api_router.post("/upload-files")
async def upload_files(files: List[UploadFile] = File(...), file_source: str = Form("Client")):
    """Upload multiple files (PDF, Excel, CSV) with source tracking"""
    try:
        uploaded_files = []
        
        # Validate file_source
        if file_source not in ["Client", "ICyte"]:
            raise HTTPException(status_code=400, detail="file_source must be 'Client' or 'ICyte'")
        
        for file in files:
            # Detect file type from MIME and extension
            content_type = file.content_type
            filename = file.filename.lower()
            
            # Determine file type tag
            if filename.endswith('.pdf'):
                file_type_tag = "PDF"
                file_type = "pdf"
            elif filename.endswith(('.xlsx', '.xls')):
                file_type_tag = "Excel"
                file_type = "excel"
            elif filename.endswith('.csv'):
                file_type_tag = "CSV"
                file_type = "csv"
            else:
                # Skip unsupported files
                continue
            
            # Validate file types for ICyte uploads (Excel/CSV only)
            if file_source == "ICyte" and file_type == "pdf":
                continue  # Skip PDFs for ICyte uploads
            
            file_id = str(uuid.uuid4())
            file_path = UPLOADS_DIR / f"{file_id}_{file.filename}"
            
            # Save file
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            # Save metadata to DB
            doc = {
                "id": file_id,
                "filename": file.filename,
                "original_filename": file.filename,
                "file_path": str(file_path),
                "file_type": file_type,
                "file_type_tag": file_type_tag,
                "content_type": content_type,
                "file_size": file_size,
                "file_source": file_source,  # Track the source: "Client" or "ICyte"
                "uploaded_at": datetime.now(timezone.utc).isoformat(),
                "uploader": "user",  # Can be extended with actual user info
                "scan_status": "passed",  # For now, auto-pass. Can integrate virus scan later
                "status": "uploaded",
                "version": 1
            }
            await db.uploads.insert_one(doc)
            # Create a clean copy without MongoDB ObjectId for response
            clean_doc = {k: v for k, v in doc.items() if k != '_id'}
            uploaded_files.append(clean_doc)
        
        return {"uploaded_files": uploaded_files, "count": len(uploaded_files)}
    except HTTPException:
        raise  # Re-raise HTTPExceptions as-is
    except Exception as e:
        logger.error(f"Upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/upload-pdf", response_model=PDFUploadResponse)
async def upload_pdf(file: UploadFile = File(...)):
    """Upload a PDF file (Legacy endpoint - use /upload-files instead)"""
    try:
        if not file.filename.endswith('.pdf'):
            raise HTTPException(status_code=400, detail="Only PDF files are allowed")
        
        file_id = str(uuid.uuid4())
        file_path = UPLOADS_DIR / f"{file_id}_{file.filename}"
        
        # Save file
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        file_size = os.path.getsize(file_path)
        
        # Save metadata to DB
        doc = {
            "id": file_id,
            "filename": file.filename,
            "original_filename": file.filename,
            "file_path": str(file_path),
            "file_type": "pdf",
            "file_type_tag": "PDF",
            "content_type": file.content_type,
            "file_size": file_size,
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "uploader": "user",
            "scan_status": "passed",
            "status": "uploaded",
            "version": 1
        }
        await db.uploads.insert_one(doc)
        
        return PDFUploadResponse(**doc)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """Upload an Excel file (Legacy endpoint - use /upload-files instead)"""
    try:
        if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files are allowed")
        
        file_id = str(uuid.uuid4())
        file_path = UPLOADS_DIR / f"{file_id}_{file.filename}"
        
        # Save file
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        file_size = os.path.getsize(file_path)
        
        # Save metadata to DB
        doc = {
            "id": file_id,
            "filename": file.filename,
            "file_path": str(file_path),
            "file_type": "excel",
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "status": "uploaded"
        }
        await db.uploads.insert_one(doc)
        
        return {"id": file_id, "filename": file.filename, "status": "uploaded"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/file/{file_id}/rename")
async def rename_file(file_id: str, new_filename: str = Form(...)):
    """Rename an uploaded file"""
    try:
        file_doc = await db.uploads.find_one({"id": file_id})
        if not file_doc:
            raise HTTPException(status_code=404, detail="File not found")
        
        # Update filename in database
        result = await db.uploads.update_one(
            {"id": file_id},
            {"$set": {"filename": new_filename}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="File not found")
        
        return {"message": "File renamed successfully", "new_filename": new_filename}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/file/{file_id}")
async def delete_file(file_id: str):
    """Delete an uploaded file"""
    try:
        file_doc = await db.uploads.find_one({"id": file_id})
        if not file_doc:
            raise HTTPException(status_code=404, detail="File not found")
        
        # Delete physical file
        file_path = file_doc['file_path']
        if os.path.exists(file_path):
            os.remove(file_path)
        
        # Delete from database
        await db.uploads.delete_one({"id": file_id})
        
        return {"message": "File deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/conversion/{conversion_id}")
async def delete_conversion(conversion_id: str):
    """Delete a converted Excel file"""
    try:
        conversion_doc = await db.conversions.find_one({"id": conversion_id})
        if not conversion_doc:
            raise HTTPException(status_code=404, detail="Conversion not found")
        
        # Delete physical Excel file
        excel_path = conversion_doc.get('excel_path')
        if excel_path and os.path.exists(excel_path):
            os.remove(excel_path)
        
        # Delete from database
        await db.conversions.delete_one({"id": conversion_id})
        
        return {"message": "Conversion deleted successfully"}
    except Exception as e:
        logger.error(f"Delete conversion error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/convert-pdf", response_model=ConversionResponse)
async def convert_pdf(request: ConversionRequest):
    """Convert PDF to Excel using LLM with custom prompt"""
    try:
        # Get file from DB
        file_doc = await db.uploads.find_one({"id": request.file_id})
        if not file_doc:
            raise HTTPException(status_code=404, detail="File not found")
        
        file_path = file_doc['file_path']
        
        # Initialize LLM Chat with Gemini (supports file attachments)
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"conversion_{request.file_id}",
            system_message="You are an expert data extraction assistant. Extract data from PDF and provide it in a clean, structured JSON format suitable for Excel conversion."
        ).with_model("gemini", "gemini-2.0-flash")
        
        # Create file attachment
        pdf_file = FileContentWithMimeType(
            file_path=file_path,
            mime_type="application/pdf"
        )
        
        # Create enhanced prompt for better extraction
        full_prompt = f"""{request.prompt}

IMPORTANT: Provide the extracted data in a clean JSON format with the following structure:
{{
    "data": [
        {{"column1": "value1", "column2": "value2", ...}},
        {{"column1": "value1", "column2": "value2", ...}}
    ]
}}

Extract ALL tables, data points, and relevant information from the PDF.
Make sure column names are clear and descriptive.
Return ONLY the JSON, no additional text or markdown formatting."""
        
        user_message = UserMessage(
            text=full_prompt,
            file_contents=[pdf_file]
        )
        
        # Get LLM response
        response = await chat.send_message(user_message)
        
        # Parse response and create Excel
        conversion_id = str(uuid.uuid4())
        excel_path = UPLOADS_DIR / f"{conversion_id}_converted.xlsx"
        
        # Try to parse JSON from response
        df = None
        json_data = None
        
        # Clean response - remove markdown code blocks if present
        cleaned_response = response.strip()
        if cleaned_response.startswith('```'):
            # Remove markdown code blocks
            lines = cleaned_response.split('\n')
            cleaned_response = '\n'.join([line for line in lines if not line.strip().startswith('```')])
        
        # Try to find and parse JSON
        try:
            # Try direct JSON parse
            json_data = json.loads(cleaned_response)
            
            if isinstance(json_data, dict) and 'data' in json_data:
                # Expected format
                df = pd.DataFrame(json_data['data'])
            elif isinstance(json_data, list):
                # Direct list format
                df = pd.DataFrame(json_data)
            elif isinstance(json_data, dict):
                # Single row or nested structure
                # Try to flatten it
                if all(isinstance(v, list) for v in json_data.values()):
                    # Column-oriented data
                    df = pd.DataFrame(json_data)
                else:
                    # Single row
                    df = pd.DataFrame([json_data])
        except json.JSONDecodeError:
            # If JSON parsing fails, try to extract structured data from text
            pass
        
        # Fallback: Create a table from the text response
        if df is None or df.empty:
            # Try to detect if response has tabular structure
            lines = cleaned_response.split('\n')
            table_data = []
            headers = None
            
            for line in lines:
                if not line.strip():
                    continue
                # Check if line looks like data (contains separators)
                if '|' in line or '\t' in line or ',' in line:
                    # Split by common separators
                    if '|' in line:
                        parts = [p.strip() for p in line.split('|') if p.strip()]
                    elif '\t' in line:
                        parts = [p.strip() for p in line.split('\t') if p.strip()]
                    else:
                        parts = [p.strip() for p in line.split(',')]
                    
                    if len(parts) >= 2:  # At least 2 columns
                        if headers is None:
                            headers = parts
                        else:
                            table_data.append(parts)
            
            if headers and table_data:
                df = pd.DataFrame(table_data, columns=headers)
            else:
                # Last resort: save as simple text output
                df = pd.DataFrame({
                    "Extracted_Data": [cleaned_response]
                })
        
        # Save to Excel
        df.to_excel(excel_path, index=False)
        
        # Save conversion to DB
        conversion_doc = {
            "id": conversion_id,
            "file_id": request.file_id,
            "excel_path": str(excel_path),
            "prompt": request.prompt,
            "status": "completed",
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.conversions.insert_one(conversion_doc)
        
        return ConversionResponse(**conversion_doc)
    except Exception as e:
        logger.error(f"Conversion error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/download-excel/{file_id}")
async def download_excel(file_id: str):
    """Download converted Excel file"""
    try:
        # Check in conversions
        conversion = await db.conversions.find_one({"id": file_id})
        if conversion:
            file_path = conversion['excel_path']
            return FileResponse(
                file_path, 
                filename=f"converted_{file_id}.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        # Check in uploads
        upload = await db.uploads.find_one({"id": file_id})
        if upload and upload.get('file_type') == 'excel':
            file_path = upload['file_path']
            return FileResponse(
                file_path, 
                filename=upload['filename'],
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        raise HTTPException(status_code=404, detail="File not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/excel-sheets/{file_id}")
async def get_excel_sheets(file_id: str):
    """Get sheet names and columns from an Excel file"""
    try:
        # Try conversion first
        conversion = await db.conversions.find_one({"id": file_id})
        if conversion:
            file_path = conversion['excel_path']
        else:
            # Try uploads
            upload = await db.uploads.find_one({"id": file_id})
            if upload and upload.get('file_type') == 'excel':
                file_path = upload['file_path']
            else:
                raise HTTPException(status_code=404, detail="File not found")
        
        # Read Excel file
        wb = openpyxl.load_workbook(file_path, data_only=False)
        sheets_info = {}
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            columns = []
            best_row_idx = None
            max_columns = 0
            
            # Try to find header row (check first 10 rows)
            # Pick the row with the most non-empty, non-formula cells
            for row_idx in range(1, min(11, sheet.max_row + 1)):
                row_cells = [cell for cell in sheet[row_idx] if cell.value is not None]
                
                # Filter out cells that are formulas or numbers (headers are usually text)
                header_like_cells = []
                for cell in row_cells:
                    val = str(cell.value)
                    # Skip if it's a formula or looks like pure numeric data
                    if not val.startswith('=') and not val.replace('.','',1).replace('-','',1).isdigit():
                        header_like_cells.append(cell)
                
                # If this row has more header-like columns, consider it
                if len(header_like_cells) > max_columns and len(header_like_cells) >= 3:
                    max_columns = len(header_like_cells)
                    best_row_idx = row_idx
                    columns = [str(cell.value).strip() for cell in row_cells]
            
            # Fallback: if no good headers found, just use first row with data
            if not columns:
                for row_idx in range(1, min(11, sheet.max_row + 1)):
                    row_values = [cell.value for cell in sheet[row_idx] if cell.value is not None]
                    if len(row_values) >= 1:
                        columns = [str(cell.value).strip() for cell in sheet[row_idx] if cell.value is not None]
                        break
            
            sheets_info[sheet_name] = columns
        
        return {"sheets": sheets_info}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Mapping Template Endpoints
@api_router.post("/mapping-templates", response_model=MappingTemplate)
async def create_mapping_template(template: MappingTemplate):
    """Create a new mapping template"""
    try:
        template_dict = template.model_dump()
        await db.mapping_templates.insert_one(template_dict)
        return template
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/mapping-templates")
async def get_mapping_templates():
    """Get all mapping templates"""
    try:
        templates = await db.mapping_templates.find({}, {"_id": 0}).to_list(100)
        return {"templates": templates}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/mapping-template/{template_id}")
async def get_mapping_template(template_id: str):
    """Get a specific mapping template"""
    try:
        template = await db.mapping_templates.find_one({"id": template_id}, {"_id": 0})
        if not template:
            raise HTTPException(status_code=404, detail="Template not found")
        return template
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/mapping-template/{template_id}", response_model=MappingTemplate)
async def update_mapping_template(template_id: str, template: MappingTemplate):
    """Update an existing mapping template"""
    try:
        template_dict = template.model_dump()
        template_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
        
        result = await db.mapping_templates.update_one(
            {"id": template_id},
            {"$set": template_dict}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Template not found")
        
        return template
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/mapping-template/{template_id}")
async def delete_mapping_template(template_id: str):
    """Delete a mapping template"""
    try:
        result = await db.mapping_templates.delete_one({"id": template_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Template not found")
        return {"message": "Template deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/configure-reconciliation", response_model=ReconciliationConfig)
async def configure_reconciliation(config: ReconciliationConfig):
    """Save reconciliation configuration"""
    try:
        config_dict = config.model_dump()
        await db.reconciliation_configs.insert_one(config_dict)
        return config
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/perform-reconciliation/{config_id}")
async def perform_reconciliation(config_id: str):
    """Perform reconciliation based on saved configuration with unique key matching"""
    try:
        # Helper function to detect header row
        def evaluate_formula(row, formula, df_columns):
            """Evaluate a formula with sequential operations
            Formula format: [
                {'column': 'Name1', 'operation': None},
                {'column': 'Name2', 'operation': 'add'},
                {'column': 'Name3', 'operation': 'subtract'}
            ]
            """
            try:
                if not formula or len(formula) == 0:
                    return None
                
                result = None
                warnings = []
                
                for idx, step in enumerate(formula):
                    col = step.get('column')
                    operation = step.get('operation')
                    
                    if not col:
                        continue
                    
                    # Check if column exists
                    if col not in df_columns:
                        warnings.append(f"Column '{col}' not found")
                        continue
                    
                    # Get value from row
                    val = row.get(col)
                    
                    # Treat NULL as 0
                    if val is None or pd.isna(val):
                        val = 0
                    
                    # Convert to numeric
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        warnings.append(f"Column '{col}' contains non-numeric value: {val}")
                        return None
                    
                    # First column - initialize result
                    if idx == 0:
                        result = val
                    else:
                        # Apply operation
                        if operation == 'add':
                            result = result + val
                        elif operation == 'subtract':
                            result = result - val
                        elif operation == 'multiply':
                            result = result * val
                        elif operation == 'divide':
                            if val == 0:
                                warnings.append(f"Division by zero: {col} = 0")
                                return None
                            result = result / val
                        else:
                            warnings.append(f"Unknown operation: {operation}")
                
                return result
            except Exception as e:
                logger.error(f"Error evaluating formula: {str(e)}")
                return None

        def find_header_row(file_path, sheet_name):
            """Find the row index where headers are located"""
            wb = openpyxl.load_workbook(file_path, data_only=False)
            sheet = wb[sheet_name]
            
            best_row_idx = 0  # Default to first row
            max_columns = 0
            
            # Check first 10 rows
            for row_idx in range(1, min(11, sheet.max_row + 1)):
                row_cells = [cell for cell in sheet[row_idx] if cell.value is not None]
                
                # Filter out cells that are formulas or numbers
                header_like_cells = []
                for cell in row_cells:
                    val = str(cell.value)
                    if not val.startswith('=') and not val.replace('.','',1).replace('-','',1).isdigit():
                        header_like_cells.append(cell)
                
                if len(header_like_cells) > max_columns and len(header_like_cells) >= 3:
                    max_columns = len(header_like_cells)
                    best_row_idx = row_idx - 1  # pandas uses 0-based indexing
            
            return best_row_idx
        
        # Get config
        config = await db.reconciliation_configs.find_one({"id": config_id})
        if not config:
            raise HTTPException(status_code=404, detail="Configuration not found")
        
        # Get client file - check conversions first, then uploads
        client_file = await db.conversions.find_one({"id": config['client_file_id']})
        if client_file:
            client_file_path = client_file['excel_path']
        else:
            # Check in uploads collection
            client_file = await db.uploads.find_one({"id": config['client_file_id']})
            if client_file:
                client_file_path = client_file['file_path']
            else:
                raise HTTPException(status_code=404, detail="Client file not found")
        
        # Get ICyte file - check uploads collection
        icyte_file = await db.uploads.find_one({"id": config['icyte_file_id']})
        if not icyte_file:
            raise HTTPException(status_code=404, detail="ICyte file not found")
        
        icyte_file_path = icyte_file['file_path']
        
        # Find header rows
        client_header_row = find_header_row(client_file_path, config['client_sheet'])
        icyte_header_row = find_header_row(icyte_file_path, config['icyte_sheet'])
        
        logger.info(f"Client header row: {client_header_row}, ICyte header row: {icyte_header_row}")
        
        # Read both Excel files with correct header rows
        client_df = pd.read_excel(client_file_path, sheet_name=config['client_sheet'], header=client_header_row)
        icyte_df = pd.read_excel(icyte_file_path, sheet_name=config['icyte_sheet'], header=icyte_header_row)
        
        # Get unique keys
        client_unique_key = config['client_unique_key']
        icyte_unique_key = config['icyte_unique_key']
        
        # Helper function to normalize NDC format (remove dashes, convert to string)
        def normalize_key(value):
            if pd.isna(value):
                return None
            # Convert to string and remove dashes and spaces
            return str(value).replace('-', '').replace(' ', '').strip().upper()
        
        # Create dictionaries for quick lookup based on unique keys
        client_dict = {}
        for idx, row in client_df.iterrows():
            key_val = row.get(client_unique_key) if client_unique_key in client_df.columns else None
            key = normalize_key(key_val)
            if key:
                client_dict[key] = row
        
        icyte_dict = {}
        for idx, row in icyte_df.iterrows():
            key_val = row.get(icyte_unique_key) if icyte_unique_key in icyte_df.columns else None
            key = normalize_key(key_val)
            if key:
                icyte_dict[key] = row
        
        # Get all unique keys from both datasets
        all_keys = set(client_dict.keys()) | set(icyte_dict.keys())
        
        # Perform reconciliation
        exceptions = []
        matched = 0
        variances = 0
        only_in_client = []
        only_in_icyte = []
        
        for key in all_keys:
            client_row = client_dict.get(key)
            icyte_row = icyte_dict.get(key)
            
            # Check if key exists in both files
            if client_row is None:
                only_in_icyte.append({
                    "unique_key": key,
                    "status": "Only in ICyte Report",
                    "details": f"{icyte_unique_key}: {key}"
                })
                variances += 1
                continue
            
            if icyte_row is None:
                only_in_client.append({
                    "unique_key": key,
                    "status": "Only in Client Report",
                    "details": f"{client_unique_key}: {key}"
                })
                variances += 1
                continue
            
            # Compare mapped columns
            row_exceptions = []
            for mapping in config['mappings']:
                # Support both old format (single columns) and new format (formula)
                client_formula = mapping.get('client_formula')
                icyte_formula = mapping.get('icyte_formula')
                
                # Fallback to old format if no formula provided
                if not client_formula:
                    client_cols = mapping.get('client_columns', [mapping.get('client_column')]) if mapping.get('client_columns') else [mapping.get('client_column')]
                    client_operation = mapping.get('client_operation', 'none')
                    # Convert old format to new formula format
                    client_formula = []
                    for i, col in enumerate(client_cols):
                        if col:
                            client_formula.append({
                                'column': col,
                                'operation': None if i == 0 else client_operation
                            })
                
                if not icyte_formula:
                    icyte_cols = mapping.get('icyte_columns', [mapping.get('icyte_column')]) if mapping.get('icyte_columns') else [mapping.get('icyte_column')]
                    icyte_operation = mapping.get('icyte_operation', 'none')
                    # Convert old format to new formula format
                    icyte_formula = []
                    for i, col in enumerate(icyte_cols):
                        if col:
                            icyte_formula.append({
                                'column': col,
                                'operation': None if i == 0 else icyte_operation
                            })
                
                # Calculate values using formula
                client_val = evaluate_formula(client_row, client_formula, client_df.columns)
                icyte_val = evaluate_formula(icyte_row, icyte_formula, icyte_df.columns)
                
                # Skip if both are NaN
                if pd.isna(client_val) and pd.isna(icyte_val):
                    continue
                
                # Operations are now handled in evaluate_formula function
                # Compare values with tolerance for floats
                values_match = False
                try:
                    if pd.isna(client_val) and pd.isna(icyte_val):
                        values_match = True
                    elif isinstance(client_val, (int, float)) and isinstance(icyte_val, (int, float)):
                        # Use tolerance for float comparison
                        values_match = abs(float(client_val) - float(icyte_val)) < 0.01
                    else:
                        values_match = str(client_val).strip() == str(icyte_val).strip()
                except:
                    values_match = str(client_val) == str(icyte_val)
                
                if not values_match:
                    # Calculate numeric variance (Client - ICyte)
                    variance_value = "mismatch"
                    try:
                        if not pd.isna(client_val) and not pd.isna(icyte_val):
                            # Handle string values with commas
                            client_str = str(client_val).replace(',', '')
                            icyte_str = str(icyte_val).replace(',', '')
                            client_num = float(client_str)
                            icyte_num = float(icyte_str)
                            variance_value = f"{client_num - icyte_num:.2f}"
                    except Exception as e:
                        logger.warning(f"Variance calculation failed for {key}: {e}")
                        variance_value = "mismatch"
                    
                    # Determine result based on variance
                    result = "Matched" if variance_value == "0.00" else "Unmatched"
                    
                    # Create column description from formula
                    client_desc = ' + '.join([step['column'] for step in client_formula if step.get('column')]) if client_formula else 'Unknown'
                    icyte_desc = ' + '.join([step['column'] for step in icyte_formula if step.get('column')]) if icyte_formula else 'Unknown'
                    
                    row_exceptions.append({
                        "unique_key": key,
                        "result": result,
                        "client_column": client_desc,
                        "icyte_column": icyte_desc,
                        "client_value": str(client_val) if not pd.isna(client_val) else "N/A",
                        "icyte_value": str(icyte_val) if not pd.isna(icyte_val) else "N/A",
                        "variance": variance_value
                    })
                    variances += 1
            
            if not row_exceptions:
                matched += 1
            else:
                exceptions.extend(row_exceptions)
        
        # Add only_in_client and only_in_icyte to exceptions
        for item in only_in_client:
            exceptions.append(item)
        for item in only_in_icyte:
            exceptions.append(item)
        
        # Create reconciliation report DataFrame for download with dynamic columns
        report_data = []
        warnings = []
        
        # Build dynamic column structure for each mapping
        for key in all_keys:
            client_row = client_dict.get(key)
            icyte_row = icyte_dict.get(key)
            
            # Base row data
            row_data = {
                config['client_unique_key']: key
            }
            
            # Determine row status
            if client_row is None:
                row_status = "MISSING_IN_CLIENT"
            elif icyte_row is None:
                row_status = "MISSING_IN_ICYTE"
            else:
                row_status = "MATCHED"  # Will be updated if variances found
            
            # Process each mapping
            all_matched = True
            for mapping in config['mappings']:
                # Support both old format (single columns) and new format (formula)
                client_formula = mapping.get('client_formula')
                icyte_formula = mapping.get('icyte_formula')
                
                # Fallback to old format if no formula provided
                if not client_formula:
                    client_col = mapping.get('client_column')
                    if client_col:
                        client_formula = [{'column': client_col, 'operation': None}]
                
                if not icyte_formula:
                    icyte_col = mapping.get('icyte_column')
                    if icyte_col:
                        icyte_formula = [{'column': icyte_col, 'operation': None}]
                
                # Calculate values using formula
                client_val = evaluate_formula(client_row, client_formula, client_df.columns) if client_row is not None else None
                icyte_val = evaluate_formula(icyte_row, icyte_formula, icyte_df.columns) if icyte_row is not None else None
                
                # Calculate variance and match flag
                variance_value = None
                match_flag = "Unmatched"
                
                if client_val is None or pd.isna(client_val) or icyte_val is None or pd.isna(icyte_val):
                    variance_value = None
                    match_flag = "Unmatched"
                else:
                    try:
                        # Keep numeric types
                        client_str = str(client_val).replace(',', '')
                        icyte_str = str(icyte_val).replace(',', '')
                        client_num = float(client_str)
                        icyte_num = float(icyte_str)
                        variance_value = client_num - icyte_num
                        
                        # Check if matched with tolerance
                        if abs(variance_value) < 0.01:
                            match_flag = "Matched"
                        else:
                            match_flag = "Unmatched"
                            all_matched = False
                    except:
                        variance_value = "N/A"
                        match_flag = "Unmatched"
                        all_matched = False
                
                # Create label from formula or custom label
                if mapping.get('label'):
                    label = mapping['label']
                elif client_formula and len(client_formula) > 0:
                    # Create label from formula columns
                    formula_cols = [step['column'] for step in client_formula if step.get('column')]
                    label = ' + '.join(formula_cols) if len(formula_cols) > 1 else formula_cols[0] if formula_cols else 'Unknown'
                else:
                    label = 'Unknown'
                
                # Add columns with proper naming convention
                # Client column with prefix
                row_data[f"Client: {label}"] = client_val if client_val is not None and not pd.isna(client_val) else None
                # ICyte column with prefix
                row_data[f"ICyte: {label}"] = icyte_val if icyte_val is not None and not pd.isna(icyte_val) else None
                # Variance (Client - ICyte) [column label]
                row_data[f"Variance (Client - ICyte) [{label}]"] = variance_value
                # Matched flag for this pair
                row_data[f"Matched [{label}]"] = match_flag
            
            # Update row status if variances found
            if client_row is not None and icyte_row is not None:
                row_status = "MATCHED" if all_matched else "VARIANCE"
            
            row_data["RowStatus"] = row_status
            report_data.append(row_data)
        
        # Save report as Excel
        report_id = str(uuid.uuid4())
        report_excel_path = UPLOADS_DIR / f"reconciliation_report_{report_id}.xlsx"
        
        # Create Excel file with all data
        if report_data:
            report_df = pd.DataFrame(report_data)
            # Preserve numeric types in Excel
            report_df.to_excel(report_excel_path, index=False)
        else:
            # Create empty report
            report_df = pd.DataFrame()
            report_df.to_excel(report_excel_path, index=False)
        
        # Count actual matched and variance records
        matched_count = len([r for r in report_data if r.get("RowStatus") == "MATCHED"])
        variance_count = len([r for r in report_data if r.get("RowStatus") == "VARIANCE"])
        
        # Build column headers info for frontend
        column_headers = {
            "unique_key": config['client_unique_key'],
            "mappings": []
        }
        for mapping in config['mappings']:
            # Support both old and new format
            client_formula = mapping.get('client_formula')
            icyte_formula = mapping.get('icyte_formula')
            
            # Fallback to old format if no formula provided
            if not client_formula:
                client_cols = mapping.get('client_columns', [mapping.get('client_column')]) if mapping.get('client_columns') else [mapping.get('client_column')]
                client_formula = [{'column': col, 'operation': None} for col in client_cols if col]
            
            if not icyte_formula:
                icyte_cols = mapping.get('icyte_columns', [mapping.get('icyte_column')]) if mapping.get('icyte_columns') else [mapping.get('icyte_column')]
                icyte_formula = [{'column': col, 'operation': None} for col in icyte_cols if col]
            
            # Create label from formula or custom label
            if mapping.get('label'):
                client_label_base = mapping['label']
                icyte_label_base = mapping['label']
            else:


@api_router.post("/save-column-mappings")
async def save_column_mappings(data: dict):
    """Save column mappings configuration"""
    try:
        mapping_id = str(uuid.uuid4())
        doc = {
            "id": mapping_id,
            "name": data.get('name', 'Untitled Mapping'),
            "client_file_id": data.get('client_file_id'),
            "icyte_file_id": data.get('icyte_file_id'),
            "client_sheet": data.get('client_sheet'),
            "icyte_sheet": data.get('icyte_sheet'),
            "mappings": data.get('mappings', []),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.column_mappings.insert_one(doc)
        return {"id": mapping_id, "message": "Mappings saved successfully"}
    except Exception as e:
        logger.error(f"Save mappings error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/upload-column-mappings")
async def upload_column_mappings(
    file: UploadFile = File(...),
    client_columns: str = Form(...),
    icyte_columns: str = Form(...)
):
    """Upload and parse column mappings from CSV file"""
    try:
        import csv
        import json
        import re
        
        # Parse available columns
        client_cols = json.loads(client_columns)
        icyte_cols = json.loads(icyte_columns)
        
        # Read CSV file
        content = await file.read()
        decoded = content.decode('utf-8')
        reader = csv.DictReader(decoded.splitlines())
        
        matched_mappings = []
        unmatched_columns = []
        
        for row in reader:
            client_expr = row.get('ClientExpression', '').strip()
            icyte_col = row.get('ICyteColumn', '').strip()
            label = row.get('Label', '').strip()
            
            if not client_expr or not icyte_col:
                continue
            
            # Parse client expression
            # Split by operators while keeping the operators
            tokens = re.split(r'(\s*[\+\-\*\/]\s*)', client_expr)
            tokens = [t.strip() for t in tokens if t.strip()]
            
            client_formula = []
            unmatched_in_expr = []
            pending_operation = None
            
            for i, token in enumerate(tokens):
                if token in ['+', '-', '*', '/']:
                    # Map operator symbols to operation names
                    op_map = {'+': 'add', '-': 'subtract', '*': 'multiply', '/': 'divide'}
                    pending_operation = op_map.get(token)
                else:
                    # This is a column name
                    if token in client_cols:
                        client_formula.append({'column': token, 'operation': pending_operation})
                        pending_operation = None
                    else:
                        unmatched_in_expr.append(('Client', token))
            
            # Check ICyte column
            if icyte_col not in icyte_cols:
                unmatched_columns.append({'side': 'ICyte', 'column': icyte_col})
            
            # Add unmatched from expression
            for side, col in unmatched_in_expr:
                unmatched_columns.append({'side': side, 'column': col})
            
            # Only add mapping if all columns matched
            if len(unmatched_in_expr) == 0 and icyte_col in icyte_cols and len(client_formula) > 0:
                matched_mappings.append({
                    'client_formula': client_formula,
                    'icyte_column': icyte_col,
                    'label': label
                })
        
        return {
            'matched_mappings': matched_mappings,
            'unmatched_columns': unmatched_columns
        }
    except Exception as e:
        logger.error(f"Upload mappings error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/column-mappings")
async def get_column_mappings():
    """Get all saved column mappings"""
    try:
        mappings = await db.column_mappings.find({}, {"_id": 0}).to_list(100)
        return {"mappings": mappings}
    except Exception as e:
        logger.error(f"Get column mappings error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/column-mapping/{mapping_id}")
async def get_column_mapping(mapping_id: str):
    """Get a specific column mapping by ID"""
    try:
        mapping = await db.column_mappings.find_one({"id": mapping_id}, {"_id": 0})
        if not mapping:
            raise HTTPException(status_code=404, detail="Column mapping not found")
        return mapping
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get column mapping error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

                    'label': label
                })
        
        return {
            "matched_mappings": matched_mappings,
            "unmatched_columns": unmatched_columns,
            "total_rows": len(matched_mappings) + len(unmatched_columns)
        }
    except Exception as e:
        logger.error(f"Upload mappings error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

                # Create label from formula columns
                client_cols = [step['column'] for step in client_formula if step.get('column')]
                icyte_cols = [step['column'] for step in icyte_formula if step.get('column')]
                client_label_base = ' + '.join(client_cols) if len(client_cols) > 1 else client_cols[0] if client_cols else 'Unknown'
                icyte_label_base = ' + '.join(icyte_cols) if len(icyte_cols) > 1 else icyte_cols[0] if icyte_cols else 'Unknown'
            
            column_headers["mappings"].append({
                "client_label": f"Client: {client_label_base}",
                "icyte_label": f"ICyte: {icyte_label_base}",
                "variance_label": f"Variance (Client - ICyte) [{client_label_base}]",
                "match_label": f"Matched [{client_label_base}]"
            })
        
        # Create report
        report = {
            "id": report_id,
            "config_id": config_id,
            "total_records": len(all_keys),
            "matched_records": matched_count,
            "variances": variance_count,
            "only_in_client": len(only_in_client),
            "only_in_icyte": len(only_in_icyte),
            "exceptions": report_data[:100],  # Limit to 100 for display
            "report_file_path": str(report_excel_path),
            "column_headers": column_headers,
            "warnings": list(set(warnings)),  # Remove duplicates
            "summary": {
                "match_rate": f"{(matched_count / len(all_keys) * 100):.2f}%" if len(all_keys) > 0 else "0%",
                "variance_rate": f"{(variance_count / len(all_keys) * 100):.2f}%" if len(all_keys) > 0 else "0%",
                "total_exceptions": len(report_data)
            },
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Insert into database
        await db.reconciliation_reports.insert_one(report.copy())
        
        # Return report without MongoDB ObjectId
        return report
    except Exception as e:
        logger.error(f"Reconciliation error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/download-reconciliation-report/{report_id}")
async def download_reconciliation_report(report_id: str):
    """Download reconciliation report as Excel"""
    try:
        report = await db.reconciliation_reports.find_one({"id": report_id}, {"_id": 0})
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")
        
        report_file_path = report.get('report_file_path')
        if not report_file_path:
            raise HTTPException(status_code=404, detail="This report was created before the download feature was added. Please create a new reconciliation to generate a downloadable report.")
        
        if os.path.exists(report_file_path):
            return FileResponse(
                report_file_path,
                filename=f"reconciliation_report_{report_id}.xlsx",
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            raise HTTPException(status_code=404, detail="Report file not found on server")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Download error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reconciliation-reports")
async def get_reconciliation_reports():
    """Get all reconciliation reports"""
    try:
        reports = await db.reconciliation_reports.find({}, {"_id": 0}).to_list(100)
        return {"reports": reports}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reconciliation-report/{report_id}")
async def get_reconciliation_report(report_id: str):
    """Get a specific reconciliation report"""
    try:
        report = await db.reconciliation_reports.find_one({"id": report_id}, {"_id": 0})
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")
        return report
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/reconciliation-report/{report_id}")
async def delete_reconciliation_report(report_id: str):
    """Delete a reconciliation report"""
    try:
        report = await db.reconciliation_reports.find_one({"id": report_id})
        if not report:
            raise HTTPException(status_code=404, detail="Report not found")
        
        # Delete physical report file if it exists
        report_file_path = report.get('report_file_path')
        if report_file_path and os.path.exists(report_file_path):
            os.remove(report_file_path)
            logger.info(f"Deleted report file: {report_file_path}")
        
        # Delete from database
        await db.reconciliation_reports.delete_one({"id": report_id})
        
        return {"message": "Report deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete report error: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/uploads")
async def get_uploads(file_source: Optional[str] = None):
    """Get all uploads, optionally filtered by file_source (Client or ICyte)"""
    try:
        query = {}
        if file_source:
            if file_source not in ["Client", "ICyte"]:
                raise HTTPException(status_code=400, detail="file_source must be 'Client' or 'ICyte'")
            query["file_source"] = file_source
        
        uploads = await db.uploads.find(query, {"_id": 0}).to_list(100)
        return {"uploads": uploads}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/conversions")
async def get_conversions():
    """Get all conversions"""
    try:
        conversions = await db.conversions.find({}, {"_id": 0}).to_list(100)
        return {"conversions": conversions}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()