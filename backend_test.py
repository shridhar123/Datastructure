#!/usr/bin/env python3
"""
Backend Test Suite for Client and ICyte File Upload Functionality
Tests separate upload functionality with file_source parameter and filtering
"""

import requests
import json
import os
import tempfile
import pandas as pd
from pathlib import Path
import time

# Get backend URL from frontend .env
def get_backend_url():
    frontend_env_path = Path("/app/frontend/.env")
    if frontend_env_path.exists():
        with open(frontend_env_path, 'r') as f:
            for line in f:
                if line.startswith('REACT_APP_BACKEND_URL='):
                    base_url = line.split('=', 1)[1].strip()
                    return f"{base_url}/api"
    return "http://localhost:8001/api"

BASE_URL = get_backend_url()
print(f"Testing backend at: {BASE_URL}")

class FileUploadTester:
    def __init__(self):
        self.test_results = []
        self.uploaded_file_ids = []
        
    def log_result(self, test_name, success, message, details=None):
        """Log test result"""
        result = {
            "test": test_name,
            "success": success,
            "message": message,
            "details": details or {}
        }
        self.test_results.append(result)
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status}: {test_name} - {message}")
        if details:
            print(f"   Details: {details}")
    
    def create_test_files(self):
        """Create sample test files for upload testing"""
        test_files = {}
        
        # Create temporary directory for test files
        temp_dir = Path(tempfile.mkdtemp())
        
        # Create sample PDF content (simple text file with .pdf extension for testing)
        pdf_path = temp_dir / "sample_invoice.pdf"
        with open(pdf_path, 'w') as f:
            f.write("Sample PDF content for testing - Invoice #12345\nAmount: $1000.00\nDate: 2024-01-15")
        test_files['pdf'] = pdf_path
        
        # Create sample Excel file
        excel_path = temp_dir / "client_data.xlsx"
        client_data = {
            'Product_ID': ['P001', 'P002', 'P003'],
            'Product_Name': ['Widget A', 'Widget B', 'Widget C'],
            'Quantity': [100, 250, 75],
            'Unit_Price': [15.99, 25.50, 8.75],
            'Total_Amount': [1599.00, 6375.00, 656.25]
        }
        df = pd.DataFrame(client_data)
        df.to_excel(excel_path, index=False)
        test_files['excel'] = excel_path
        
        # Create sample CSV file
        csv_path = temp_dir / "icyte_report.csv"
        icyte_data = {
            'NDC': ['12345-678-90', '23456-789-01', '34567-890-12'],
            'Drug_Name': ['Medication A', 'Medication B', 'Medication C'],
            'Quantity_Dispensed': [120, 240, 90],
            'Unit_Cost': [12.50, 18.75, 22.00],
            'Total_Cost': [1500.00, 4500.00, 1980.00]
        }
        df_csv = pd.DataFrame(icyte_data)
        df_csv.to_csv(csv_path, index=False)
        test_files['csv'] = csv_path
        
        return test_files
    
    def test_client_file_uploads(self, test_files):
        """Test uploading Client files (PDF, Excel, CSV) - all should succeed"""
        print("\n=== Testing Client File Uploads ===")
        
        for file_type, file_path in test_files.items():
            try:
                with open(file_path, 'rb') as f:
                    files = {'files': (file_path.name, f, self.get_mime_type(file_type))}
                    data = {'file_source': 'Client'}
                    
                    response = requests.post(f"{BASE_URL}/upload-files", files=files, data=data)
                    
                    if response.status_code == 200:
                        result = response.json()
                        uploaded_files = result.get('uploaded_files', [])
                        
                        if uploaded_files:
                            file_info = uploaded_files[0]
                            self.uploaded_file_ids.append(file_info['id'])
                            
                            # Verify file_source is set correctly
                            if file_info.get('file_source') == 'Client':
                                self.log_result(
                                    f"Client {file_type.upper()} Upload",
                                    True,
                                    f"Successfully uploaded {file_type} file with Client source",
                                    {
                                        "file_id": file_info['id'],
                                        "filename": file_info['filename'],
                                        "file_source": file_info['file_source'],
                                        "file_type": file_info['file_type']
                                    }
                                )
                            else:
                                self.log_result(
                                    f"Client {file_type.upper()} Upload",
                                    False,
                                    f"File uploaded but file_source incorrect: {file_info.get('file_source')}",
                                    file_info
                                )
                        else:
                            self.log_result(
                                f"Client {file_type.upper()} Upload",
                                False,
                                "No files returned in response",
                                result
                            )
                    else:
                        self.log_result(
                            f"Client {file_type.upper()} Upload",
                            False,
                            f"Upload failed with status {response.status_code}",
                            {"response": response.text}
                        )
                        
            except Exception as e:
                self.log_result(
                    f"Client {file_type.upper()} Upload",
                    False,
                    f"Exception during upload: {str(e)}"
                )
    
    def test_icyte_file_uploads(self, test_files):
        """Test uploading ICyte files (Excel/CSV should succeed, PDF should be skipped)"""
        print("\n=== Testing ICyte File Uploads ===")
        
        for file_type, file_path in test_files.items():
            try:
                with open(file_path, 'rb') as f:
                    files = {'files': (file_path.name, f, self.get_mime_type(file_type))}
                    data = {'file_source': 'ICyte'}
                    
                    response = requests.post(f"{BASE_URL}/upload-files", files=files, data=data)
                    
                    if response.status_code == 200:
                        result = response.json()
                        uploaded_files = result.get('uploaded_files', [])
                        
                        if file_type == 'pdf':
                            # PDF should be skipped for ICyte uploads
                            if len(uploaded_files) == 0:
                                self.log_result(
                                    f"ICyte {file_type.upper()} Upload (Skip Test)",
                                    True,
                                    "PDF correctly skipped for ICyte upload",
                                    {"count": result.get('count', 0)}
                                )
                            else:
                                self.log_result(
                                    f"ICyte {file_type.upper()} Upload (Skip Test)",
                                    False,
                                    "PDF was not skipped - should have been rejected",
                                    result
                                )
                        else:
                            # Excel and CSV should succeed
                            if uploaded_files:
                                file_info = uploaded_files[0]
                                self.uploaded_file_ids.append(file_info['id'])
                                
                                if file_info.get('file_source') == 'ICyte':
                                    self.log_result(
                                        f"ICyte {file_type.upper()} Upload",
                                        True,
                                        f"Successfully uploaded {file_type} file with ICyte source",
                                        {
                                            "file_id": file_info['id'],
                                            "filename": file_info['filename'],
                                            "file_source": file_info['file_source'],
                                            "file_type": file_info['file_type']
                                        }
                                    )
                                else:
                                    self.log_result(
                                        f"ICyte {file_type.upper()} Upload",
                                        False,
                                        f"File uploaded but file_source incorrect: {file_info.get('file_source')}",
                                        file_info
                                    )
                            else:
                                self.log_result(
                                    f"ICyte {file_type.upper()} Upload",
                                    False,
                                    f"{file_type} file should have been uploaded but wasn't",
                                    result
                                )
                    else:
                        self.log_result(
                            f"ICyte {file_type.upper()} Upload",
                            False,
                            f"Upload failed with status {response.status_code}",
                            {"response": response.text}
                        )
                        
            except Exception as e:
                self.log_result(
                    f"ICyte {file_type.upper()} Upload",
                    False,
                    f"Exception during upload: {str(e)}"
                )
    
    def test_file_filtering(self):
        """Test filtering uploads by file_source"""
        print("\n=== Testing File Source Filtering ===")
        
        # Test getting all files
        try:
            response = requests.get(f"{BASE_URL}/uploads")
            if response.status_code == 200:
                all_files = response.json().get('uploads', [])
                self.log_result(
                    "Get All Uploads",
                    True,
                    f"Retrieved {len(all_files)} total files",
                    {"count": len(all_files)}
                )
            else:
                self.log_result(
                    "Get All Uploads",
                    False,
                    f"Failed to get all uploads: {response.status_code}",
                    {"response": response.text}
                )
        except Exception as e:
            self.log_result(
                "Get All Uploads",
                False,
                f"Exception getting all uploads: {str(e)}"
            )
        
        # Test filtering by Client
        try:
            response = requests.get(f"{BASE_URL}/uploads?file_source=Client")
            if response.status_code == 200:
                client_files = response.json().get('uploads', [])
                
                # Verify all returned files have file_source=Client
                all_client = all(f.get('file_source') == 'Client' for f in client_files)
                
                if all_client:
                    self.log_result(
                        "Filter Client Files",
                        True,
                        f"Retrieved {len(client_files)} Client files, all correctly filtered",
                        {"count": len(client_files), "files": [f['filename'] for f in client_files]}
                    )
                else:
                    non_client = [f for f in client_files if f.get('file_source') != 'Client']
                    self.log_result(
                        "Filter Client Files",
                        False,
                        f"Found {len(non_client)} non-Client files in Client filter",
                        {"non_client_files": non_client}
                    )
            else:
                self.log_result(
                    "Filter Client Files",
                    False,
                    f"Failed to filter Client files: {response.status_code}",
                    {"response": response.text}
                )
        except Exception as e:
            self.log_result(
                "Filter Client Files",
                False,
                f"Exception filtering Client files: {str(e)}"
            )
        
        # Test filtering by ICyte
        try:
            response = requests.get(f"{BASE_URL}/uploads?file_source=ICyte")
            if response.status_code == 200:
                icyte_files = response.json().get('uploads', [])
                
                # Verify all returned files have file_source=ICyte
                all_icyte = all(f.get('file_source') == 'ICyte' for f in icyte_files)
                
                if all_icyte:
                    self.log_result(
                        "Filter ICyte Files",
                        True,
                        f"Retrieved {len(icyte_files)} ICyte files, all correctly filtered",
                        {"count": len(icyte_files), "files": [f['filename'] for f in icyte_files]}
                    )
                else:
                    non_icyte = [f for f in icyte_files if f.get('file_source') != 'ICyte']
                    self.log_result(
                        "Filter ICyte Files",
                        False,
                        f"Found {len(non_icyte)} non-ICyte files in ICyte filter",
                        {"non_icyte_files": non_icyte}
                    )
            else:
                self.log_result(
                    "Filter ICyte Files",
                    False,
                    f"Failed to filter ICyte files: {response.status_code}",
                    {"response": response.text}
                )
        except Exception as e:
            self.log_result(
                "Filter ICyte Files",
                False,
                f"Exception filtering ICyte files: {str(e)}"
            )
    
    def test_file_metadata_verification(self):
        """Test that uploaded files have correct metadata"""
        print("\n=== Testing File Metadata Verification ===")
        
        if not self.uploaded_file_ids:
            self.log_result(
                "Metadata Verification",
                False,
                "No uploaded files to verify metadata"
            )
            return
        
        # Get all uploads and verify metadata for our uploaded files
        try:
            response = requests.get(f"{BASE_URL}/uploads")
            if response.status_code == 200:
                all_files = response.json().get('uploads', [])
                our_files = [f for f in all_files if f['id'] in self.uploaded_file_ids]
                
                metadata_issues = []
                for file_info in our_files:
                    # Check required metadata fields
                    required_fields = ['id', 'filename', 'file_source', 'file_type', 'file_type_tag', 'file_size']
                    missing_fields = [field for field in required_fields if field not in file_info]
                    
                    if missing_fields:
                        metadata_issues.append(f"File {file_info['id']}: Missing fields {missing_fields}")
                    
                    # Verify file_source is valid
                    if file_info.get('file_source') not in ['Client', 'ICyte']:
                        metadata_issues.append(f"File {file_info['id']}: Invalid file_source '{file_info.get('file_source')}'")
                    
                    # Verify file_type matches file_type_tag
                    file_type = file_info.get('file_type')
                    file_type_tag = file_info.get('file_type_tag')
                    expected_tags = {'pdf': 'PDF', 'excel': 'Excel', 'csv': 'CSV'}
                    
                    if file_type in expected_tags and file_type_tag != expected_tags[file_type]:
                        metadata_issues.append(f"File {file_info['id']}: file_type_tag mismatch - expected {expected_tags[file_type]}, got {file_type_tag}")
                
                if not metadata_issues:
                    self.log_result(
                        "File Metadata Verification",
                        True,
                        f"All {len(our_files)} uploaded files have correct metadata",
                        {"verified_files": len(our_files)}
                    )
                else:
                    self.log_result(
                        "File Metadata Verification",
                        False,
                        f"Found {len(metadata_issues)} metadata issues",
                        {"issues": metadata_issues}
                    )
            else:
                self.log_result(
                    "File Metadata Verification",
                    False,
                    f"Failed to get uploads for metadata verification: {response.status_code}"
                )
        except Exception as e:
            self.log_result(
                "File Metadata Verification",
                False,
                f"Exception during metadata verification: {str(e)}"
            )
    
    def test_invalid_file_source(self):
        """Test that invalid file_source values are rejected"""
        print("\n=== Testing Invalid File Source Validation ===")
        
        try:
            # Create a simple test file
            temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False)
            temp_file.write("test content")
            temp_file.close()
            
            with open(temp_file.name, 'rb') as f:
                files = {'files': ('test.txt', f, 'text/plain')}
                data = {'file_source': 'InvalidSource'}
                
                response = requests.post(f"{BASE_URL}/upload-files", files=files, data=data)
                
                if response.status_code == 400:
                    self.log_result(
                        "Invalid File Source Validation",
                        True,
                        "Invalid file_source correctly rejected with 400 status",
                        {"status_code": response.status_code}
                    )
                else:
                    self.log_result(
                        "Invalid File Source Validation",
                        False,
                        f"Invalid file_source not rejected - got status {response.status_code}",
                        {"response": response.text}
                    )
            
            # Clean up
            os.unlink(temp_file.name)
            
        except Exception as e:
            self.log_result(
                "Invalid File Source Validation",
                False,
                f"Exception testing invalid file_source: {str(e)}"
            )
    
    def get_mime_type(self, file_type):
        """Get MIME type for file type"""
        mime_types = {
            'pdf': 'application/pdf',
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'csv': 'text/csv'
        }
        return mime_types.get(file_type, 'application/octet-stream')
    
    def cleanup_test_files(self):
        """Clean up uploaded test files"""
        print("\n=== Cleaning Up Test Files ===")
        
        for file_id in self.uploaded_file_ids:
            try:
                response = requests.delete(f"{BASE_URL}/file/{file_id}")
                if response.status_code == 200:
                    print(f"✅ Cleaned up file {file_id}")
                else:
                    print(f"⚠️  Failed to clean up file {file_id}: {response.status_code}")
            except Exception as e:
                print(f"⚠️  Exception cleaning up file {file_id}: {str(e)}")
    
    def run_all_tests(self):
        """Run all tests"""
        print("🚀 Starting Backend Tests for Client and ICyte File Upload Functionality")
        print("=" * 80)
        
        # Create test files
        test_files = self.create_test_files()
        print(f"📁 Created test files: {list(test_files.keys())}")
        
        # Run tests
        self.test_client_file_uploads(test_files)
        self.test_icyte_file_uploads(test_files)
        self.test_file_filtering()
        self.test_file_metadata_verification()
        self.test_invalid_file_source()
        
        # Print summary
        self.print_summary()
        
        # Cleanup
        self.cleanup_test_files()
        
        return self.test_results
    
    def print_summary(self):
        """Print test summary"""
        print("\n" + "=" * 80)
        print("📊 TEST SUMMARY")
        print("=" * 80)
        
        passed = sum(1 for result in self.test_results if result['success'])
        failed = len(self.test_results) - passed
        
        print(f"Total Tests: {len(self.test_results)}")
        print(f"✅ Passed: {passed}")
        print(f"❌ Failed: {failed}")
        print(f"Success Rate: {(passed/len(self.test_results)*100):.1f}%")
        
        if failed > 0:
            print("\n🔍 FAILED TESTS:")
            for result in self.test_results:
                if not result['success']:
                    print(f"  ❌ {result['test']}: {result['message']}")
        
        print("\n" + "=" * 80)

class ReconciliationWorkflowTester:
    def __init__(self):
        self.test_results = []
        self.client_file_id = None
        self.icyte_file_id = None
        self.config_id = None
        self.report_id = None
        
    def log_result(self, test_name, success, message, details=None):
        """Log test result"""
        result = {
            "test": test_name,
            "success": success,
            "message": message,
            "details": details or {}
        }
        self.test_results.append(result)
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status}: {test_name} - {message}")
        if details:
            print(f"   Details: {details}")
    
    def create_reconciliation_test_files(self):
        """Create test files specifically for reconciliation testing"""
        test_files = {}
        temp_dir = Path(tempfile.mkdtemp())
        
        # Create Client Excel file with NDC data
        client_excel_path = temp_dir / "client_reconciliation_data.xlsx"
        client_data = {
            'NDC11': ['12345678901', '23456789012', '34567890123', '45678901234'],
            'Drug_Name': ['Aspirin 325mg', 'Ibuprofen 200mg', 'Acetaminophen 500mg', 'Naproxen 220mg'],
            'Quantity': [100, 250, 150, 75],
            'Unit_Price': [0.15, 0.25, 0.20, 0.35],
            'Total_Amount': [15.00, 62.50, 30.00, 26.25],
            'Manufacturer': ['Bayer', 'Advil', 'Tylenol', 'Aleve']
        }
        df_client = pd.DataFrame(client_data)
        df_client.to_excel(client_excel_path, index=False, sheet_name='ClientData')
        test_files['client_excel'] = client_excel_path
        
        # Create ICyte Excel file with matching and non-matching NDC data
        icyte_excel_path = temp_dir / "icyte_reconciliation_report.xlsx"
        icyte_data = {
            'NDC_Code': ['12345678901', '23456789012', '34567890123', '56789012345'],  # Last one is different
            'Product_Name': ['Aspirin 325mg', 'Ibuprofen 200mg', 'Acetaminophen 500mg', 'Omeprazole 20mg'],
            'Dispensed_Qty': [100, 250, 140, 60],  # Third one has variance (150 vs 140)
            'Cost_Per_Unit': [0.15, 0.25, 0.22, 0.45],  # Third one has variance (0.20 vs 0.22)
            'Total_Cost': [15.00, 62.50, 30.80, 27.00],
            'Supplier': ['Bayer Corp', 'Advil Inc', 'Tylenol LLC', 'Prilosec Co']
        }
        df_icyte = pd.DataFrame(icyte_data)
        df_icyte.to_excel(icyte_excel_path, index=False, sheet_name='ICyteReport')
        test_files['icyte_excel'] = icyte_excel_path
        
        return test_files
    
    def upload_reconciliation_files(self, test_files):
        """Upload test files for reconciliation"""
        print("\n=== Uploading Reconciliation Test Files ===")
        
        # Upload Client file
        try:
            with open(test_files['client_excel'], 'rb') as f:
                files = {'files': (test_files['client_excel'].name, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {'file_source': 'Client'}
                
                response = requests.post(f"{BASE_URL}/upload-files", files=files, data=data)
                
                if response.status_code == 200:
                    result = response.json()
                    uploaded_files = result.get('uploaded_files', [])
                    if uploaded_files:
                        self.client_file_id = uploaded_files[0]['id']
                        self.log_result(
                            "Upload Client Reconciliation File",
                            True,
                            f"Successfully uploaded client file: {uploaded_files[0]['filename']}",
                            {"file_id": self.client_file_id}
                        )
                    else:
                        self.log_result("Upload Client Reconciliation File", False, "No files returned")
                        return False
                else:
                    self.log_result("Upload Client Reconciliation File", False, f"Upload failed: {response.status_code}")
                    return False
        except Exception as e:
            self.log_result("Upload Client Reconciliation File", False, f"Exception: {str(e)}")
            return False
        
        # Upload ICyte file
        try:
            with open(test_files['icyte_excel'], 'rb') as f:
                files = {'files': (test_files['icyte_excel'].name, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                data = {'file_source': 'ICyte'}
                
                response = requests.post(f"{BASE_URL}/upload-files", files=files, data=data)
                
                if response.status_code == 200:
                    result = response.json()
                    uploaded_files = result.get('uploaded_files', [])
                    if uploaded_files:
                        self.icyte_file_id = uploaded_files[0]['id']
                        self.log_result(
                            "Upload ICyte Reconciliation File",
                            True,
                            f"Successfully uploaded ICyte file: {uploaded_files[0]['filename']}",
                            {"file_id": self.icyte_file_id}
                        )
                    else:
                        self.log_result("Upload ICyte Reconciliation File", False, "No files returned")
                        return False
                else:
                    self.log_result("Upload ICyte Reconciliation File", False, f"Upload failed: {response.status_code}")
                    return False
        except Exception as e:
            self.log_result("Upload ICyte Reconciliation File", False, f"Exception: {str(e)}")
            return False
        
        return True
    
    def test_get_available_files(self):
        """Test Step 1: Get available files"""
        print("\n=== Testing Get Available Files ===")
        
        # Test GET /api/conversions
        try:
            response = requests.get(f"{BASE_URL}/conversions")
            if response.status_code == 200:
                conversions = response.json().get('conversions', [])
                self.log_result(
                    "Get Conversions",
                    True,
                    f"Retrieved {len(conversions)} conversion files",
                    {"count": len(conversions)}
                )
            else:
                self.log_result("Get Conversions", False, f"Failed: {response.status_code}")
        except Exception as e:
            self.log_result("Get Conversions", False, f"Exception: {str(e)}")
        
        # Test GET /api/uploads?file_source=Client
        try:
            response = requests.get(f"{BASE_URL}/uploads?file_source=Client")
            if response.status_code == 200:
                client_files = response.json().get('uploads', [])
                self.log_result(
                    "Get Client Files",
                    True,
                    f"Retrieved {len(client_files)} Client files",
                    {"count": len(client_files), "files": [f['filename'] for f in client_files]}
                )
            else:
                self.log_result("Get Client Files", False, f"Failed: {response.status_code}")
        except Exception as e:
            self.log_result("Get Client Files", False, f"Exception: {str(e)}")
        
        # Test GET /api/uploads?file_source=ICyte
        try:
            response = requests.get(f"{BASE_URL}/uploads?file_source=ICyte")
            if response.status_code == 200:
                icyte_files = response.json().get('uploads', [])
                self.log_result(
                    "Get ICyte Files",
                    True,
                    f"Retrieved {len(icyte_files)} ICyte files",
                    {"count": len(icyte_files), "files": [f['filename'] for f in icyte_files]}
                )
            else:
                self.log_result("Get ICyte Files", False, f"Failed: {response.status_code}")
        except Exception as e:
            self.log_result("Get ICyte Files", False, f"Exception: {str(e)}")
    
    def test_get_excel_sheets(self):
        """Test Step 2: Get Excel sheet information"""
        print("\n=== Testing Get Excel Sheets ===")
        
        if not self.client_file_id or not self.icyte_file_id:
            self.log_result("Get Excel Sheets", False, "Missing file IDs for testing")
            return False
        
        # Test Client file sheets
        try:
            response = requests.get(f"{BASE_URL}/excel-sheets/{self.client_file_id}")
            if response.status_code == 200:
                sheets_info = response.json().get('sheets', {})
                if sheets_info:
                    self.log_result(
                        "Get Client Excel Sheets",
                        True,
                        f"Retrieved sheets and columns for client file",
                        {"sheets": list(sheets_info.keys()), "columns": sheets_info}
                    )
                else:
                    self.log_result("Get Client Excel Sheets", False, "No sheets returned")
                    return False
            else:
                self.log_result("Get Client Excel Sheets", False, f"Failed: {response.status_code}")
                return False
        except Exception as e:
            self.log_result("Get Client Excel Sheets", False, f"Exception: {str(e)}")
            return False
        
        # Test ICyte file sheets
        try:
            response = requests.get(f"{BASE_URL}/excel-sheets/{self.icyte_file_id}")
            if response.status_code == 200:
                sheets_info = response.json().get('sheets', {})
                if sheets_info:
                    self.log_result(
                        "Get ICyte Excel Sheets",
                        True,
                        f"Retrieved sheets and columns for ICyte file",
                        {"sheets": list(sheets_info.keys()), "columns": sheets_info}
                    )
                else:
                    self.log_result("Get ICyte Excel Sheets", False, "No sheets returned")
                    return False
            else:
                self.log_result("Get ICyte Excel Sheets", False, f"Failed: {response.status_code}")
                return False
        except Exception as e:
            self.log_result("Get ICyte Excel Sheets", False, f"Exception: {str(e)}")
            return False
        
        return True
    
    def test_configure_reconciliation(self):
        """Test Step 3: Configure reconciliation"""
        print("\n=== Testing Configure Reconciliation ===")
        
        if not self.client_file_id or not self.icyte_file_id:
            self.log_result("Configure Reconciliation", False, "Missing file IDs for configuration")
            return False
        
        # Create reconciliation configuration
        config_data = {
            "client_file_id": self.client_file_id,
            "icyte_file_id": self.icyte_file_id,
            "client_sheet": "ClientData",
            "icyte_sheet": "ICyteReport",
            "client_unique_key": "NDC11",
            "icyte_unique_key": "NDC_Code",
            "mappings": [
                {
                    "client_column": "Quantity",
                    "icyte_column": "Dispensed_Qty"
                },
                {
                    "client_column": "Unit_Price",
                    "icyte_column": "Cost_Per_Unit"
                },
                {
                    "client_column": "Total_Amount",
                    "icyte_column": "Total_Cost"
                }
            ]
        }
        
        try:
            response = requests.post(f"{BASE_URL}/configure-reconciliation", json=config_data)
            if response.status_code == 200:
                config_result = response.json()
                self.config_id = config_result.get('id')
                if self.config_id:
                    self.log_result(
                        "Configure Reconciliation",
                        True,
                        f"Successfully created reconciliation configuration",
                        {"config_id": self.config_id, "mappings": len(config_data['mappings'])}
                    )
                    return True
                else:
                    self.log_result("Configure Reconciliation", False, "No config ID returned")
                    return False
            else:
                self.log_result("Configure Reconciliation", False, f"Failed: {response.status_code} - {response.text}")
                return False
        except Exception as e:
            self.log_result("Configure Reconciliation", False, f"Exception: {str(e)}")
            return False
    
    def test_perform_reconciliation(self):
        """Test Step 4: Perform reconciliation"""
        print("\n=== Testing Perform Reconciliation ===")
        
        if not self.config_id:
            self.log_result("Perform Reconciliation", False, "Missing config ID for reconciliation")
            return False
        
        try:
            response = requests.post(f"{BASE_URL}/perform-reconciliation/{self.config_id}")
            if response.status_code == 200:
                report_result = response.json()
                self.report_id = report_result.get('id')
                
                # Verify response structure
                required_fields = ['id', 'total_records', 'matched_records', 'variances', 'column_headers', 'warnings', 'exceptions']
                missing_fields = [field for field in required_fields if field not in report_result]
                
                if missing_fields:
                    self.log_result(
                        "Perform Reconciliation",
                        False,
                        f"Missing required fields in response: {missing_fields}",
                        report_result
                    )
                    return False
                
                # Verify numeric values are preserved
                total_records = report_result.get('total_records')
                matched_records = report_result.get('matched_records')
                variances = report_result.get('variances')
                
                if not isinstance(total_records, int) or not isinstance(matched_records, int) or not isinstance(variances, int):
                    self.log_result(
                        "Perform Reconciliation",
                        False,
                        "Numeric values not preserved correctly",
                        {"total_records": type(total_records), "matched_records": type(matched_records), "variances": type(variances)}
                    )
                    return False
                
                # Verify column headers structure
                column_headers = report_result.get('column_headers', {})
                if 'unique_key' not in column_headers or 'mappings' not in column_headers:
                    self.log_result(
                        "Perform Reconciliation",
                        False,
                        "Column headers structure incorrect",
                        column_headers
                    )
                    return False
                
                # Verify exceptions data structure
                exceptions = report_result.get('exceptions', [])
                if exceptions and isinstance(exceptions, list):
                    # Check first exception for proper column structure
                    first_exception = exceptions[0]
                    expected_columns = ['NDC11', 'RowStatus']  # Unique key and row status
                    
                    # Should also have dynamic columns for mappings
                    for mapping in column_headers.get('mappings', []):
                        expected_columns.extend([
                            mapping.get('client_label'),
                            mapping.get('icyte_label'),
                            mapping.get('variance_label'),
                            mapping.get('match_label')
                        ])
                    
                    # Check if some expected columns exist
                    found_columns = [col for col in expected_columns if col in first_exception]
                    
                    if len(found_columns) < 3:  # At least unique key + some mapping columns
                        self.log_result(
                            "Perform Reconciliation",
                            False,
                            f"Exception data missing expected columns. Found: {list(first_exception.keys())}",
                            {"expected_some_of": expected_columns, "found": list(first_exception.keys())}
                        )
                        return False
                
                self.log_result(
                    "Perform Reconciliation",
                    True,
                    f"Successfully performed reconciliation",
                    {
                        "report_id": self.report_id,
                        "total_records": total_records,
                        "matched_records": matched_records,
                        "variances": variances,
                        "warnings_count": len(report_result.get('warnings', [])),
                        "exceptions_count": len(exceptions)
                    }
                )
                return True
            else:
                self.log_result("Perform Reconciliation", False, f"Failed: {response.status_code} - {response.text}")
                return False
        except Exception as e:
            self.log_result("Perform Reconciliation", False, f"Exception: {str(e)}")
            return False
    
    def test_verify_report_structure(self):
        """Test Step 5: Verify report structure"""
        print("\n=== Testing Verify Report Structure ===")
        
        if not self.report_id:
            self.log_result("Verify Report Structure", False, "Missing report ID for verification")
            return False
        
        try:
            response = requests.get(f"{BASE_URL}/reconciliation-report/{self.report_id}")
            if response.status_code == 200:
                report = response.json()
                
                # Verify dynamic columns structure
                exceptions = report.get('exceptions', [])
                if not exceptions:
                    self.log_result(
                        "Verify Report Structure",
                        True,
                        "Report structure verified (no exceptions to check)",
                        {"report_id": self.report_id}
                    )
                    return True
                
                first_row = exceptions[0]
                
                # Check for unique key column
                if 'NDC11' not in first_row:
                    self.log_result("Verify Report Structure", False, "Missing unique key column (NDC11)")
                    return False
                
                # Check for RowStatus column
                if 'RowStatus' not in first_row:
                    self.log_result("Verify Report Structure", False, "Missing RowStatus column")
                    return False
                
                # Verify RowStatus values
                valid_statuses = ['MATCHED', 'VARIANCE', 'MISSING_IN_CLIENT', 'MISSING_IN_ICYTE']
                row_status = first_row.get('RowStatus')
                if row_status not in valid_statuses:
                    self.log_result(
                        "Verify Report Structure",
                        False,
                        f"Invalid RowStatus value: {row_status}",
                        {"valid_statuses": valid_statuses}
                    )
                    return False
                
                # Check for mapping columns (Client:, ICyte:, Variance, Matched)
                mapping_columns = [col for col in first_row.keys() if 
                                 col.startswith('Client:') or 
                                 col.startswith('ICyte:') or 
                                 col.startswith('Variance') or 
                                 col.startswith('Matched')]
                
                if len(mapping_columns) < 6:  # Should have at least 2 columns per mapping * 3 mappings
                    self.log_result(
                        "Verify Report Structure",
                        False,
                        f"Insufficient mapping columns found: {len(mapping_columns)}",
                        {"found_columns": mapping_columns}
                    )
                    return False
                
                # Verify numeric values are preserved (not strings)
                numeric_columns = [col for col in first_row.keys() if 'Client:' in col or 'ICyte:' in col or 'Variance' in col]
                numeric_issues = []
                for col in numeric_columns:
                    value = first_row.get(col)
                    if value is not None and col != 'Matched [Quantity]' and col != 'Matched [Unit_Price]' and col != 'Matched [Total_Amount]':
                        # Skip 'Matched' columns as they contain strings
                        if 'Variance' not in col and isinstance(value, str) and value.replace('.', '').replace('-', '').isdigit():
                            numeric_issues.append(f"{col}: {value} (should be numeric, got string)")
                
                if numeric_issues:
                    self.log_result(
                        "Verify Report Structure",
                        False,
                        f"Numeric values not preserved: {len(numeric_issues)} issues",
                        {"issues": numeric_issues[:3]}  # Show first 3 issues
                    )
                    return False
                
                self.log_result(
                    "Verify Report Structure",
                    True,
                    "Report structure verified successfully",
                    {
                        "unique_key_present": True,
                        "row_status_valid": True,
                        "mapping_columns_count": len(mapping_columns),
                        "numeric_values_preserved": True
                    }
                )
                return True
            else:
                self.log_result("Verify Report Structure", False, f"Failed to get report: {response.status_code}")
                return False
        except Exception as e:
            self.log_result("Verify Report Structure", False, f"Exception: {str(e)}")
            return False
    
    def test_fetch_reports(self):
        """Test Step 6: Fetch reports"""
        print("\n=== Testing Fetch Reports ===")
        
        # Test GET /api/reconciliation-reports
        try:
            response = requests.get(f"{BASE_URL}/reconciliation-reports")
            if response.status_code == 200:
                reports_list = response.json().get('reports', [])
                
                # Check if our report appears in the list
                our_report = None
                for report in reports_list:
                    if report.get('id') == self.report_id:
                        our_report = report
                        break
                
                if our_report:
                    self.log_result(
                        "Fetch Reports List",
                        True,
                        f"Report appears in list (total: {len(reports_list)} reports)",
                        {"our_report_id": self.report_id, "total_reports": len(reports_list)}
                    )
                else:
                    self.log_result(
                        "Fetch Reports List",
                        False,
                        f"Our report not found in list of {len(reports_list)} reports",
                        {"report_ids": [r.get('id') for r in reports_list]}
                    )
            else:
                self.log_result("Fetch Reports List", False, f"Failed: {response.status_code}")
        except Exception as e:
            self.log_result("Fetch Reports List", False, f"Exception: {str(e)}")
        
        # Test GET /api/reconciliation-report/{report_id}
        if self.report_id:
            try:
                response = requests.get(f"{BASE_URL}/reconciliation-report/{self.report_id}")
                if response.status_code == 200:
                    report_data = response.json()
                    
                    # Verify full report data is accessible
                    required_fields = ['id', 'config_id', 'total_records', 'matched_records', 'variances', 'exceptions']
                    missing_fields = [field for field in required_fields if field not in report_data]
                    
                    if not missing_fields:
                        self.log_result(
                            "Fetch Individual Report",
                            True,
                            "Full report data accessible",
                            {"report_id": self.report_id, "fields_present": len(required_fields)}
                        )
                    else:
                        self.log_result(
                            "Fetch Individual Report",
                            False,
                            f"Missing fields in report data: {missing_fields}",
                            report_data
                        )
                else:
                    self.log_result("Fetch Individual Report", False, f"Failed: {response.status_code}")
            except Exception as e:
                self.log_result("Fetch Individual Report", False, f"Exception: {str(e)}")
    
    def cleanup_reconciliation_files(self):
        """Clean up test files"""
        print("\n=== Cleaning Up Reconciliation Test Files ===")
        
        for file_id, file_type in [(self.client_file_id, "Client"), (self.icyte_file_id, "ICyte")]:
            if file_id:
                try:
                    response = requests.delete(f"{BASE_URL}/file/{file_id}")
                    if response.status_code == 200:
                        print(f"✅ Cleaned up {file_type} file {file_id}")
                    else:
                        print(f"⚠️  Failed to clean up {file_type} file {file_id}: {response.status_code}")
                except Exception as e:
                    print(f"⚠️  Exception cleaning up {file_type} file {file_id}: {str(e)}")
    
    def run_reconciliation_workflow_tests(self):
        """Run complete reconciliation workflow tests"""
        print("🚀 Starting Reconciliation Workflow Tests")
        print("=" * 80)
        
        # Create and upload test files
        test_files = self.create_reconciliation_test_files()
        print(f"📁 Created reconciliation test files: {list(test_files.keys())}")
        
        if not self.upload_reconciliation_files(test_files):
            print("❌ Failed to upload test files - aborting reconciliation tests")
            return self.test_results
        
        # Run workflow tests
        self.test_get_available_files()
        
        if self.test_get_excel_sheets():
            if self.test_configure_reconciliation():
                if self.test_perform_reconciliation():
                    self.test_verify_report_structure()
                    self.test_fetch_reports()
        
        # Print summary
        self.print_summary()
        
        # Cleanup
        self.cleanup_reconciliation_files()
        
        return self.test_results
    
    def print_summary(self):
        """Print test summary"""
        print("\n" + "=" * 80)
        print("📊 RECONCILIATION WORKFLOW TEST SUMMARY")
        print("=" * 80)
        
        passed = sum(1 for result in self.test_results if result['success'])
        failed = len(self.test_results) - passed
        
        print(f"Total Tests: {len(self.test_results)}")
        print(f"✅ Passed: {passed}")
        print(f"❌ Failed: {failed}")
        print(f"Success Rate: {(passed/len(self.test_results)*100):.1f}%")
        
        if failed > 0:
            print("\n🔍 FAILED TESTS:")
            for result in self.test_results:
                if not result['success']:
                    print(f"  ❌ {result['test']}: {result['message']}")
        
        print("\n" + "=" * 80)

class DownloadFunctionalityTester:
    def __init__(self):
        self.test_results = []
        self.conversion_id = None
        
    def log_result(self, test_name, success, message, details=None):
        """Log test result"""
        result = {
            "test": test_name,
            "success": success,
            "message": message,
            "details": details or {}
        }
        self.test_results.append(result)
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status}: {test_name} - {message}")
        if details:
            print(f"   Details: {details}")
    
    def test_get_conversions_list(self):
        """Test Step 1: Get conversions list and find completed conversions"""
        print("\n=== Testing Get Conversions List ===")
        
        try:
            response = requests.get(f"{BASE_URL}/conversions")
            if response.status_code == 200:
                conversions = response.json().get('conversions', [])
                
                # Look for conversions with status="completed"
                completed_conversions = [c for c in conversions if c.get('status') == 'completed']
                
                if completed_conversions:
                    # Pick the first completed conversion
                    self.conversion_id = completed_conversions[0]['id']
                    self.log_result(
                        "Get Conversions List",
                        True,
                        f"Found {len(completed_conversions)} completed conversions out of {len(conversions)} total",
                        {
                            "total_conversions": len(conversions),
                            "completed_conversions": len(completed_conversions),
                            "selected_conversion_id": self.conversion_id,
                            "selected_conversion": completed_conversions[0]
                        }
                    )
                    return True
                else:
                    self.log_result(
                        "Get Conversions List",
                        False,
                        f"No completed conversions found out of {len(conversions)} total conversions",
                        {"conversions": conversions}
                    )
                    return False
            else:
                self.log_result(
                    "Get Conversions List",
                    False,
                    f"Failed to get conversions: {response.status_code}",
                    {"response": response.text}
                )
                return False
        except Exception as e:
            self.log_result(
                "Get Conversions List",
                False,
                f"Exception getting conversions: {str(e)}"
            )
            return False
    
    def test_download_excel_endpoint(self):
        """Test Step 2: Test download endpoint with valid conversion ID"""
        print("\n=== Testing Download Excel Endpoint ===")
        
        if not self.conversion_id:
            self.log_result(
                "Download Excel Endpoint",
                False,
                "No conversion ID available for testing"
            )
            return False
        
        try:
            response = requests.get(f"{BASE_URL}/download-excel/{self.conversion_id}")
            
            if response.status_code == 200:
                # Check Content-Type header
                content_type = response.headers.get('content-type', '')
                expected_content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                
                content_type_valid = content_type == expected_content_type
                
                # Check Content-Disposition header
                content_disposition = response.headers.get('content-disposition', '')
                has_filename = 'filename=' in content_disposition
                
                # Check file content (Excel files start with PK signature)
                content = response.content
                file_size = len(content)
                is_excel_file = content.startswith(b'PK')  # Excel files are ZIP-based and start with PK
                
                # Verify all conditions
                all_checks_passed = (
                    content_type_valid and
                    has_filename and
                    file_size > 0 and
                    is_excel_file
                )
                
                if all_checks_passed:
                    self.log_result(
                        "Download Excel Endpoint",
                        True,
                        f"Successfully downloaded valid Excel file ({file_size} bytes)",
                        {
                            "conversion_id": self.conversion_id,
                            "content_type": content_type,
                            "content_disposition": content_disposition,
                            "file_size": file_size,
                            "is_excel_format": is_excel_file,
                            "headers_valid": True
                        }
                    )
                    return True
                else:
                    issues = []
                    if not content_type_valid:
                        issues.append(f"Invalid Content-Type: expected '{expected_content_type}', got '{content_type}'")
                    if not has_filename:
                        issues.append(f"Missing filename in Content-Disposition header: '{content_disposition}'")
                    if file_size == 0:
                        issues.append("File size is 0 bytes")
                    if not is_excel_file:
                        issues.append("File does not have Excel signature (PK)")
                    
                    self.log_result(
                        "Download Excel Endpoint",
                        False,
                        f"Download validation failed: {len(issues)} issues found",
                        {
                            "issues": issues,
                            "content_type": content_type,
                            "content_disposition": content_disposition,
                            "file_size": file_size,
                            "is_excel_format": is_excel_file
                        }
                    )
                    return False
            else:
                self.log_result(
                    "Download Excel Endpoint",
                    False,
                    f"Download failed with status {response.status_code}",
                    {
                        "conversion_id": self.conversion_id,
                        "status_code": response.status_code,
                        "response": response.text
                    }
                )
                return False
        except Exception as e:
            self.log_result(
                "Download Excel Endpoint",
                False,
                f"Exception during download: {str(e)}"
            )
            return False
    
    def test_download_error_cases(self):
        """Test Step 3: Test error cases with invalid conversion ID"""
        print("\n=== Testing Download Error Cases ===")
        
        # Test with invalid conversion ID
        invalid_id = "invalid-conversion-id-12345"
        
        try:
            response = requests.get(f"{BASE_URL}/download-excel/{invalid_id}")
            
            if response.status_code == 404:
                self.log_result(
                    "Download Error Cases - Invalid ID",
                    True,
                    f"Correctly returned 404 for invalid conversion ID",
                    {
                        "invalid_id": invalid_id,
                        "status_code": response.status_code,
                        "response": response.text
                    }
                )
            else:
                self.log_result(
                    "Download Error Cases - Invalid ID",
                    False,
                    f"Expected 404 for invalid ID, got {response.status_code}",
                    {
                        "invalid_id": invalid_id,
                        "status_code": response.status_code,
                        "response": response.text
                    }
                )
        except Exception as e:
            self.log_result(
                "Download Error Cases - Invalid ID",
                False,
                f"Exception testing invalid ID: {str(e)}"
            )
        
        # Test with empty conversion ID
        try:
            response = requests.get(f"{BASE_URL}/download-excel/")
            
            # This should return 404 or 405 (Method Not Allowed) depending on routing
            if response.status_code in [404, 405]:
                self.log_result(
                    "Download Error Cases - Empty ID",
                    True,
                    f"Correctly handled empty conversion ID with status {response.status_code}",
                    {
                        "status_code": response.status_code
                    }
                )
            else:
                self.log_result(
                    "Download Error Cases - Empty ID",
                    False,
                    f"Unexpected status for empty ID: {response.status_code}",
                    {
                        "status_code": response.status_code,
                        "response": response.text
                    }
                )
        except Exception as e:
            self.log_result(
                "Download Error Cases - Empty ID",
                False,
                f"Exception testing empty ID: {str(e)}"
            )
    
    def test_file_save_and_open(self):
        """Test Step 4: Test that downloaded file can be saved and opened"""
        print("\n=== Testing File Save and Open ===")
        
        if not self.conversion_id:
            self.log_result(
                "File Save and Open",
                False,
                "No conversion ID available for testing"
            )
            return False
        
        try:
            response = requests.get(f"{BASE_URL}/download-excel/{self.conversion_id}")
            
            if response.status_code == 200:
                # Save file to temporary location
                import tempfile
                import openpyxl
                
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                    temp_file.write(response.content)
                    temp_file_path = temp_file.name
                
                try:
                    # Try to open the file with openpyxl to verify it's a valid Excel file
                    workbook = openpyxl.load_workbook(temp_file_path)
                    sheet_names = workbook.sheetnames
                    
                    # Get some basic info about the file
                    first_sheet = workbook[sheet_names[0]] if sheet_names else None
                    row_count = first_sheet.max_row if first_sheet else 0
                    col_count = first_sheet.max_column if first_sheet else 0
                    
                    workbook.close()
                    
                    self.log_result(
                        "File Save and Open",
                        True,
                        f"Successfully saved and opened Excel file",
                        {
                            "conversion_id": self.conversion_id,
                            "file_path": temp_file_path,
                            "sheet_count": len(sheet_names),
                            "sheet_names": sheet_names,
                            "first_sheet_rows": row_count,
                            "first_sheet_cols": col_count
                        }
                    )
                    
                    # Clean up temp file
                    os.unlink(temp_file_path)
                    return True
                    
                except Exception as excel_error:
                    self.log_result(
                        "File Save and Open",
                        False,
                        f"File saved but could not be opened as Excel: {str(excel_error)}",
                        {
                            "conversion_id": self.conversion_id,
                            "file_path": temp_file_path,
                            "excel_error": str(excel_error)
                        }
                    )
                    # Clean up temp file
                    try:
                        os.unlink(temp_file_path)
                    except:
                        pass
                    return False
            else:
                self.log_result(
                    "File Save and Open",
                    False,
                    f"Could not download file for save test: {response.status_code}",
                    {"conversion_id": self.conversion_id}
                )
                return False
        except Exception as e:
            self.log_result(
                "File Save and Open",
                False,
                f"Exception during file save and open test: {str(e)}"
            )
            return False
    
    def run_download_tests(self):
        """Run all download functionality tests"""
        print("🚀 Starting Download Functionality Tests")
        print("=" * 80)
        
        # Run tests in sequence
        if self.test_get_conversions_list():
            self.test_download_excel_endpoint()
            self.test_file_save_and_open()
        
        # Always test error cases
        self.test_download_error_cases()
        
        # Print summary
        self.print_summary()
        
        return self.test_results
    
    def print_summary(self):
        """Print test summary"""
        print("\n" + "=" * 80)
        print("📊 DOWNLOAD FUNCTIONALITY TEST SUMMARY")
        print("=" * 80)
        
        passed = sum(1 for result in self.test_results if result['success'])
        failed = len(self.test_results) - passed
        
        print(f"Total Tests: {len(self.test_results)}")
        print(f"✅ Passed: {passed}")
        print(f"❌ Failed: {failed}")
        print(f"Success Rate: {(passed/len(self.test_results)*100):.1f}%")
        
        if failed > 0:
            print("\n🔍 FAILED TESTS:")
            for result in self.test_results:
                if not result['success']:
                    print(f"  ❌ {result['test']}: {result['message']}")
        
        print("\n" + "=" * 80)

if __name__ == "__main__":
    # Run download functionality tests (as requested)
    print("Running Download Functionality Tests...")
    download_tester = DownloadFunctionalityTester()
    download_results = download_tester.run_download_tests()
    
    print("\n" + "="*100 + "\n")
    
    # Run file upload tests
    print("Running File Upload Tests...")
    upload_tester = FileUploadTester()
    upload_results = upload_tester.run_all_tests()
    
    print("\n" + "="*100 + "\n")
    
    # Run reconciliation workflow tests
    print("Running Reconciliation Workflow Tests...")
    reconciliation_tester = ReconciliationWorkflowTester()
    reconciliation_results = reconciliation_tester.run_reconciliation_workflow_tests()
    
    # Combined summary
    print("\n" + "="*100)
    print("🎯 COMBINED TEST SUMMARY")
    print("="*100)
    
    total_download_tests = len(download_results)
    passed_download_tests = sum(1 for result in download_results if result['success'])
    
    total_upload_tests = len(upload_results)
    passed_upload_tests = sum(1 for result in upload_results if result['success'])
    
    total_reconciliation_tests = len(reconciliation_results)
    passed_reconciliation_tests = sum(1 for result in reconciliation_results if result['success'])
    
    total_tests = total_download_tests + total_upload_tests + total_reconciliation_tests
    total_passed = passed_download_tests + passed_upload_tests + passed_reconciliation_tests
    
    print(f"Download Tests: {passed_download_tests}/{total_download_tests} passed")
    print(f"File Upload Tests: {passed_upload_tests}/{total_upload_tests} passed")
    print(f"Reconciliation Tests: {passed_reconciliation_tests}/{total_reconciliation_tests} passed")
    print(f"Overall: {total_passed}/{total_tests} passed ({(total_passed/total_tests*100):.1f}%)")
    print("="*100)